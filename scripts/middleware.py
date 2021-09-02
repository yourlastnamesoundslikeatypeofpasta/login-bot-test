import logging
import os
import datetime

import requests
import openpyxl

from scripts.production_score import get_production_score
from scripts.validate_input import validate_input
from scripts.mistakes import Mistakes
from scripts.send_mistake_report import MistakeReport
from pprint import pprint


def fetch_user(payload, context, next):
    user = payload["user"]
    context['user'] = user
    next()


def fetch_trigger_id(body, context, event, next):
    trigger_id = body['trigger_id']
    context['trigger_id'] = trigger_id
    next()


def is_points_clear_block(body, context, next):
    try:
        points_clear_block = body['view']['blocks'][5]
        if points_clear_block:
            context['is_points_clear_block'] = True
            context['points'] = body['view']['blocks'][5]['text']['text']
            next()
    except IndexError:
        context['is_points_clear_block'] = False
        next()
    """values_lst = []
    package_input = body['view']['state']['values']['block_package']['package_input']['value']
    item_input = body['view']['state']['values']['block_items']['item_input']['value']
    weight_input = body['view']['state']['values']['block_weight']['weight_input']['value']
    tier_input = body['view']['state']['values']['block_tier']['static_tier_selected_do_nothing_please'][
        'selected_option']
    values_lst.append(package_input)
    values_lst.append(item_input)
    values_lst.append(weight_input)
    values_lst.append(tier_input)
    for value in values_lst:
        if value is None:
            context['private_metadata'] = 'True'
            next()
            return
    next()
"""


def calculate_production_score(view, context, next):
    # values entered
    input_block_values = {
        "block_package": view['state']['values']['block_package']['package_input']['value'].strip(' '),
        "block_weight": view['state']['values']['block_weight']['weight_input']['value'].strip(' '),
        "block_items": view['state']['values']['block_items']['item_input']['value'].strip(' '),
        "block_hours": view['state']['values']['block_hours']['hour_input']['value'].strip(' ')
    }

    # validate input block values
    error_response_action = validate_input(input_block_values)
    if error_response_action:
        # throw validation error if blocks don't meet input req
        context['error_response'] = error_response_action
        next()
        return

    # add stats to context
    context['package_count'] = float(input_block_values['block_package'])
    context['weight_count'] = float(input_block_values['block_weight'])
    context['item_count'] = float(input_block_values['block_items'])
    context['hour_count'] = float(input_block_values['block_hours'])
    context['pkg_per_hour'] = context['package_count'] / context['hour_count']
    context['weight_per_package'] = context['weight_count'] / context['package_count']
    context['items_per_pkg'] = context['item_count'] / context['package_count']
    context['production_score'] = get_production_score(context['package_count'],
                                                       context['weight_count'],
                                                       context['item_count'],
                                                       context['hour_count'])
    next()


def fetch_current_mistake_points(body, context, next):
    context['mistake_points'] = body['view']['private_metadata']
    next()


def fetch_selected_mistake_points(body, context, next):
    selected_mistake_points = body['view']['state']['values']['block_mistake_static_select']['action_static_mistake']['selected_option']['value']
    if context['mistake_points']:
        total_points = int(context['mistake_points']) + int(selected_mistake_points)
    else:
        total_points = selected_mistake_points

    context['mistake_points'] = str(total_points)
    next()


""" mistake_code = 
    body['view']['state']['values']['block_mistake_static_select']['action_static_mistake']['selected_option'][
        'value']

context['root_view_id'] = body['view']['root_view_id']
# check if a Mistakes instance exists
if are_inputs_empty:  # todo: mistakes are not disappearing when modal is closed and reopened
    # create one
    mistakes = Mistakes()
    mistakes.add_mistake(mistake_code)
    context['mistake_points'] = mistakes.get_mistake_points()
    pprint(Mistakes.instances)
    next()
    return
else:
    # use the last instance
    mistakes = Mistakes.instances[-1]
mistakes.add_mistake(mistake_code)
context['mistake_points'] = mistakes.get_mistake_points()
next()"""


def fetch_root_id(body, context, next):
    context['root_view_id'] = body['view']['root_view_id']
    next()


def download_file_shared(body, context, event, next, logger):
    # import is here because if added at the top, it breaks other middleware funcs, not sure why
    # todo: find out why import app from main breaks other funcs
    from main import app

    # get file info
    file_id = body['event']['files'][0]['id']
    info = app.client.files_info(file=file_id)
    file_name = info['file']['name']
    file_download_link = info['file']['url_private_download']

    # download file to path
    token = os.environ['bot_token']
    headers = {
        'Authorization': f"Bearer {token}"
    }  # token in headers is needed to download private files from the workspace
    r = requests.get(url=file_download_link, headers=headers)
    file_downloads_dir = os.path.join('resources', 'downloads')
    file_download_path = os.path.join('resources', 'downloads', file_name)
    if not os.path.exists(file_downloads_dir):
        os.makedirs(file_download_path)

    try:
        with open(file_download_path, 'wb') as f:
            for chunk in r.iter_content():
                f.write(chunk)
    except PermissionError:
        logger.error('Permissions Error')
    mistakes = MistakeReport(file_download_path)
    next()


def parse_file_download(body, context, next, logger):
    mistake_report = MistakeReport.instances[-1]

    file_download_path = mistake_report.file_download_path
    # open workbook and assign sheets
    mistake_report_file = file_download_path

    mistake_report_workbook = openpyxl.load_workbook(mistake_report_file)

    mistake_report_workbook_names = mistake_report_workbook.sheetnames
    sheet1, sheet2 = mistake_report_workbook[mistake_report_workbook_names[0]], mistake_report_workbook[
        mistake_report_workbook_names[1]]

    # close wb, values saved in sheet[number]
    mistake_report_workbook.close()

    # collects a list of the name on the mistake report, will find out how to perform code block via .max_row method
    num_rows_mistake_names = 0
    for i in range(0, 1048576):  # calculates how many rows of data in column 1
        logger = sheet2.cell(row=i + 3, column=1).value  # goes through each cell in column 1
        if logger is not None:  # todo: fix this old crap code
            num_rows_mistake_names += 1
            continue
        break

    # creates a sorted(set()) list of logger names that have mistakes, located in {loggerName}sheet (A**x T***a)
    mistake_names = sorted((list(set([sheet2.cell(row=i + 3, column=1).value for i in range(num_rows_mistake_names)]))))

    # checks to see if any loggers with a channel have a mistake
    # paper counts how many pieces of paper were saved
    employee_mistake_lst = []
    paper = 0
    for employee in mistake_names:
        # find mistakes employee made
        mistake_type = [sheet2.cell(row=i + 3, column=4).value  # list of mistakes associated with employee
                        for i in range(num_rows_mistake_names)
                        if sheet2.cell(row=i + 3, column=1).value == employee]

        incident_date = [sheet2.cell(row=i + 3, column=3).value  # list of incident dates associated with employee
                         for i in range(num_rows_mistake_names)
                         if sheet2.cell(row=i + 3, column=1).value == employee]

        entered_date = [sheet2.cell(row=i + 3, column=2).value  # list of entered dates associated with employee
                        for i in range(num_rows_mistake_names)
                        if sheet2.cell(row=i + 3, column=1).value == employee]

        suite = [sheet2.cell(row=i + 3, column=5).value  # list of suites associated with employee
                 for i in range(num_rows_mistake_names)
                 if sheet2.cell(row=i + 3, column=1).value == employee]

        pkg_id = [sheet2.cell(row=i + 3, column=6).value  # list of pkg ids associated with employee
                  for i in range(num_rows_mistake_names)
                  if sheet2.cell(row=i + 3, column=1).value == employee]

        incident_notes = [sheet2.cell(row=i + 3, column=8).value  # list of incident notes associated with employee
                          for i in range(num_rows_mistake_names)
                          if sheet2.cell(row=i + 3, column=1).value == employee]

        formatted_incident_date_lst = []
        formatted_entered_date_lst = []
        # format dates
        for date in incident_date:
            formatted_date = date.strftime('%m/%d/%y')
            formatted_incident_date_lst.append(formatted_date)
        for date in entered_date:
            formatted_date = date.strftime('%m/%d/%y')
            formatted_entered_date_lst.append(formatted_date)

        mistake_lst_zipped = list(zip(formatted_entered_date_lst, formatted_incident_date_lst,
                                      mistake_type, suite, pkg_id, incident_notes))
        mistake_lst = []
        for mistake in mistake_lst_zipped:
            employee_mistake_dict = {
                "entered_date": mistake[0],
                "incident_date": mistake[1],
                "mistake_type": mistake[2],
                "suite": mistake[3],
                "pkg_id": mistake[4],
                "incident_notes": mistake[5]
            }
            mistake_lst.append(employee_mistake_dict)

        employee_dict = {
            "employee_name": employee,
            "employee_mistakes": mistake_lst
        }
        employee_mistake_lst.append(employee_dict)
        paper += 1
    context['paper'] = paper
    context['employee_mistake_dict'] = employee_mistake_lst
    next()
