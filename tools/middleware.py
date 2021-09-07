import logging
import os
import datetime

import requests
import openpyxl

from tools.send_mistake_report import MistakeReport


def download_file_shared(body, context, event, message, next, logger):
    # get file info
    client = context['client']
    file_id = body['event']['files'][0]['id']
    info = client.files_info(file=file_id)
    file_name = info['file']['name']
    file_download_link = info['file']['url_private_download']

    # download file to path
    token = os.environ['bot_token']
    headers = {
        'Authorization': f"Bearer {token}"}  # token in headers is needed to download private files from the workspace
    r = requests.get(url=file_download_link, headers=headers)
    file_downloads_dir = os.path.join('data')
    file_download_path = os.path.join('data', file_name)
    if not os.path.exists(file_downloads_dir):
        os.makedirs(file_download_path)

    try:
        with open(file_download_path, 'wb') as f:
            for chunk in r.iter_content():
                f.write(chunk)
    except PermissionError:
        logger.error('Permissions Error: File is most likely open in Excel. Please close Excel and restart')
    mistakes = MistakeReport(file_download_path)
    next()


def parse_file_download(body, context, next, logger):
    # open workbook, assign sheets, close wb
    mistake_report_obj = MistakeReport.instances[-1]
    file_download_path = mistake_report_obj.file_download_path
    mistake_report_file = file_download_path
    wb = openpyxl.load_workbook(mistake_report_file)
    wb_sheet_lst = wb.sheetnames
    sheet = wb[wb_sheet_lst[1]]
    wb.close()

    # create a list of dicts with lists for each column associated  with the employee
    num_rows_w_mistakes = sheet.max_row
    mistake_report = {}
    for row in list(sheet.iter_rows(3, num_rows_w_mistakes)):
        mistake = {}
        employee = None
        for cell in row:
            if cell.column_letter == 'A':
                employee = cell.value
            elif cell.column_letter == 'B':
                entered_date = cell.value.strftime('%m/%d/%y')
                mistake['entered_date'] = entered_date
            elif cell.column_letter == 'C':
                incident_date = cell.value.strftime('%m/%d/%y')
                mistake['incident_date'] = incident_date
            elif cell.column_letter == 'D':
                mistake_type = cell.value
                mistake['mistake_type'] = mistake_type
            elif cell.column_letter == 'E':
                suite = cell.value
                mistake['suite'] = suite
            elif cell.column_letter == 'F':
                pkg_id = cell.value
                mistake['pkg_id'] = pkg_id
            elif cell.column_letter == 'H':
                incident_notes = cell.value
                mistake['incident_notes'] = incident_notes
            else:
                continue

        if employee in mistake_report:
            mistake_report[employee].append(mistake)
        else:
            mistake_report[employee] = [mistake]
        context['mistake_report_dict'] = mistake_report
    next()
